"""Goal: extract key data from sample workbook (my favorite foods) and import it into my personal worksheet"""

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from collections import namedtuple

wb = load_workbook('hello.xlsx')
ws = wb.active

categories = ['Category', 'FoodItem', 'Measure', 'Calories', 'Protein', 'Fat', 'Carbs', 'Fiber']
# this named tuple will let me add the associated values of each food item into its corresponding category easily.
Categories = namedtuple('Categories', categories)

Nutella = Categories('Sweets', 'Spread, chocolate hazelnut', '30mL', 203, 2, 11, 23, 2.0)  # Set of nutella's data

n = 0
for col in range(1, 9):  # cells A1-H1 are filled with the values from the 'categories' list
    char = get_column_letter(col)
    ws[char + '1'].value = categories[n]
    n += 1

# Create a workbook object
# wb = Workbook()


wb.save('hello.xlsx')  # saves workbook changes